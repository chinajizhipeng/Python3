#coding=utf-8
import 模块#导入XX模块
model.fun() #这样调用~防止函数混淆
from 模块 import 功能 #从model模块导入fun函数
fun() #这样调用
from model import fun as fun1 #从model模块导入fun函数,并命名为fun1
from model as m #model模块并命名为m
from model * #导入所有函数
dir(model) #查看属性，返回model包含的函数
model_name_ #模块的名字
from imp import reload
reload(module) #重载模块
help([object])
type(object) #给出对象的类型
print("",end = "") #打印不换行
#可变类型：列表、字典
#不可变类型：数字、字符串、元组
a+=a 不等于 a=a+a 前者对a直接进行修改，左边为重新定义引用  数字看上去是一样
##模块
import random
random.random() #生成0-1的一个数
random.uniform(3,4) #生成指定范围的一个数
random.randint(4,7) #生成指定范围的一个整数
random.choice()  #是从序列中获取一个随机元素，这个序列可以是字符串，元组，列表等
random.shuffle()   #是将一个列表中的元素打乱
random.sample()    #是从指定序列中随机获取指定长度的片断
x = random.choice(['hello,world','hello,jizhipeng']) #从序列中随机选择一个元素
random.randint(a,b) #生成给定值之间的一个随机数
import datetime
datetime.date.today()        #今天的日期
datetime.date.today().day    #今天的日期的日
datetime.date.today().month  #今天的日期的月
datetime.date.today().year   #今天的日期的年
datetime.date.isoformat(datetime.date.today()) #将日期转换成字符串
datetime.striptime(aru,format) #传入实参，更改时间的格式 #《从入门到实践 317页》
import time  #查看运行时间
start_time = time.time()
code
end = time.time()
time.sleep(1) #隔一秒再运行一次
#将date对象转换为字符串
import datetime
import time
oct21st = datetime.datetime(2018,7,25,9,45,0)
oct21st.strftime('%Y/%m/%d %H:%M:%S') #'2018/07/25 09:45:00'
oct21st.strftime('%I:%M %P')
#将字符串转换为datetime对象
datetime.datetime.striptime('October 21,2015','%B %b,%Y')

from collections import OrdereDict #创建字典并记录其中的项的添加顺序
a  = orderDict()
a['a1']='b1' #有顺序

import sys
sys.path #包的路径
sys.path.append("path") #添加路径
#用户输入
name = input()
name = raw_input() #py2中的input，将输入的当成代码运行
a = "x"
a += "/ny"
name = input(a)  #当提示超过一行时
int() #获取数值输入，将字符串表示转换为数值表示
#运算符号：
nan #not a number
/ #除法
// #整除
% #余数，求模运算符
** #幂运算
pow(2,3) #2的3次方
x = 3 #赋值
int() #取整
abs() #绝对值
round() #最接近的整数
range(0,10) #0-9
list(range(0,10)) #
import math #导入math模块
math.floor() #向下取整
math.ceil() #向上取整
##编码
#decode的作用是将其他编码的字符串转换成 Unicode 编码
#encode的作用是将 Unicode 编码转换成其他编码的字符串
#一句话：UTF-8是对Unicode字符集进行编码的一种编码方式
#任何平台的任何编码 都能和 Unicode 互相转换UTF-8 与 GBK 互相转换，那就先把UTF-8转换成Unicode，再从Unicode转换成GBK，
s = '我爱python' #utf-8和gbk编码一一对应
s1 = s.encode()
s1.decode()
s2 = s.encode('gbk')
s2.decode('gbk')

###列表和元组
[]  #序列
a = 'hello'
a[0]  #h
a[-1] #o
'hello'[1] #e
#切片 不回改变列表的状态
for i in d:
    enumerate(i) # i在d里的偏移量
tag = 'abcdef'
tag[2:3]   #c
tag[-3:-1] 
tag[0:3][1]
tag[-3:] 
tag[:3]
tag[:] #整个序列
tag[::2] #每两步选择一次
tag[0:10:2] #步长为2
'a'.join(['v','b','f']) #vabaf 将列表转换为字符串
len() #返回序列包含的元素个数
max() #序列包含的最大个数
min()
sum() #列表相加
for i,j in zip([L1,l2]): #对序列进行打包
for i in reversed(L1): #序列逆序
map(fun,list) #以参数序列中的每一个元素调用函数，返回包含函数返回值的新序列 PY3返回迭代器
list(filter(lambda x: x > 0,range(-5,5))) #筛选大于0的值
list(zip(X,Y)) #要用list表示出来

from functools import reduce
reduce((lambda x,y: x+y),[1,2,3,4])
#列表生成式
[x for x in range(1,18)] #后面的for...只控制循环的次数
[x for x in range(1,18) if x %2 == 0]
[(x,y) for x in range(3) for y in range(2)] #[(0, 0), (0, 1), (1, 0), (1, 1), (2, 0), (2, 1)]

#成员资格
in
a = 'rw'
'w' in a #返回布尔值
##列表 会改变列表的状态
list() #将序列转换为列表
x = [1,1,1]
x[1] = 2 #修改为[1，2，1]
del x[1] #删除2 变为[1,1]
object.method(argument)
x.append(4) #变为1，1，1，4 #append加单个元素 添加整体
a.clear() #清除列表
a = b #ab同时指向一个副本，一个变，另一个也变
b = a.copy() #将b变为a的副本
x.count(1) #1出现的次数
a.extend(x) #将x附加到a的末尾 整体当作单个元素添加到其中
x.index(1) #1第一次出现的索引
x.insert(1,'2') #在x的第1个位置插入2
x.pop(0)    #删除第0个元素，并返回这一元素 默认删除最后一个
x.remove()  #删除第一个指定的元素,不是索引值
x.reverse() #按相反的顺序排列，永久性修改
list(reversed(a)) #生成新列表对象
x.sort()    #对原来的列表进行修改,排序
y = sorted(x) #获取排序后列表的副本
x.sort(key=len) #根据长度对元素进行排序
x.sort(key=str.lower) #按照字母顺序排列 a A z Z
x.sort(reverse=True) #按相反方向进行排序
#给切片赋值
a = list('pell')
a[2:0] = list('aa') #变为peaa
##元组
#元组的不变心只适用于元组本身顶层而并非其内容 嵌套里层可修改
(1,2,3)
a,b,c = (1,2,3) #元组可以拆包
(42,) #只包含一个元素时，要让元组真正成为元组，再小括号之间至少要包含一个逗号
tuple(list) #将列表转换为元组
list(tuple) #将元组转化为列表
a1.index(2) #搜索
a1.count(2) #计数
##字符串
{}
\t #制表符
\n #换行符
print("\n")
"{foo} {} {bar} {}".format(1,2,bar=4,foo=3) #'3 1 4 2'
"{foo} {0} {bar} {1}".format(1,2,bar=4,foo=3) #'3 2 4 1' 添加索引
"the number is {num:b}".format(num = 42) #字符串格式设置中的类型说明符 。PY基础教程45
"asd%s,asdas%d".%(name,num)
#设置宽度
"{num:10}".format(num=3) #设置宽度为10 '         3'
"a is {pi:.2f}".format(pi=0.333333333333) #设置精度 小数点后2位
"a is {pi:10.2f}".format(pi=0.333333333333) #同时设定 pi的宽度的精度
"a is {:,}".format(10**100) #逗号设置千位分隔符
"{:010.2f}".format(0.333333333333)
print("{0:<10.2f}\n{0:^10.2f}\n{0:>10.2f}".format(3333.33)) #<左对齐 >右对齐 =居中
"{:$^15}".format(" WIN ") #$符号填充 '$$$$$ WIN $$$$$'
"{:#g.2f}".format(42) #必须保留小数点
#字符串方法
"hello".center(15)   #两边填充字符居中
"hello".center(15,"!") #'!!!!!hello!!!!!'
str1.rjust(2)
str1.ljust(2)
"hello".find('e') #e出现的第一个索引，没有时返回-1
"hello".rfind('e') #从右边找
"hello".index('e') #没有时返回异常
"hello".rindex('e') #从右边找
"hello".find('l',2，3) #l从第2个索引到第3个索引开始数
str1.endwith("") #以结尾
str1.startwith("") #以开头
name.upper() #将字符串全部大些
name.lower() #将字符串全部小写
Str.partition("word") #word之前分开，之后分开，变成三部分 返回元组
Str.splitlines() #按照/n进行切割 每一行的内容
'sssssssssss'.title() #大写第一个字母
import string 
string.capwords("ssssssss") #大写第一个字母
" a s d f g".replace("a","A") #将a替换为A，可将空格替换
str1.replace("a","A",1) #将a替换为A,替换超过不超过几次
"a/s/d/f".split('/') #['a', 's', 'd', 'f']  有空格 \t \n时直接用split
"a/s/d/f".split('-')
"    a     ".strip() #将开头和末尾的空格删除，不包括中间
a.rstrip() #末尾多余的空白删除
a.lstrip() #删除开头的空白
"!*!*!a !*!*!".strip("!*") #将开头和末尾的!*删除，不包括中间
str() #将非字符串值表示为字符串
table = str.maketrans("cs","kz") #c对应k s对应z
"csccscscscscccsc".translate(table) #translate进行点对点的转换
##字典
a_a = {"a":"1","b":"2"}
d = dict(a_a) #创建字典。
d = dict{"a":42,"name":"G"} #使用关键字实参来调用这个函数
d["a"]  # "1"
del.d["a"] #删除d字典中的a键
lend(d) #项数
d[k] #返回与键k相关的值
d[k] = v #将值v关联到键k
del d[k] #删除键k的项
k in d #检查字典d是否包含键k的项
"aaaaaaa is {a}.".format_map(d) #'aaaaaaa is 1.'
d.clear() #删除所有字典项
a = d.copy() #返回新字典。浅复制，修改原件发生变化；替换原件不变。
import copy 
a1=copy.deepcopy(d) #深拷贝.原本修改，不改变。
a1 = copy.copy(d) #当d为可变类型时只深拷贝第一层引用 ，之后的层数识别不了。当d为不可变类型时直接为浅拷贝
{}.fromkeys([]) #创建新字典 键对应的值为NONE
dict.fromkeys(["A","B"],"b") #键A键B值为b
d.get('name') #若没有返回None
d.get('key',value) #若没有返回value
d.item() #返回一个包含所有字典项的列表
d.keys() #返回键
list(d.keys())
d.pop() #返回与指定键相关联的值，并将项从字典中删除
d.popitem() #随机返回项并删除 
d.setdefault("A","a") #指定键不存在时返回a,存在时返回对应值
d.update(x) #将x中的项更新到d中
d.values() #返回值组成的字典视图，可能保存重复值
dict(zip(['a','b'],[1,2])) #{'a': 1, 'b': 2}
for k,v in sorted(d.items())
for k in d.keys(): #遍历所有的键
for v in d.values:
for v in set(d.values): #不要重复值
dict1.sort(key = lambda x:x['name']) #对字典中的name排序
#字典中的字典
a = {}
a['s'] = {'name': 'ji'}
import pprint
pprint.pprint(a) #美观打印
a['s']['name'] #访问复杂字典
#集合
a = {'n','a'}  #返回时没有重复值
set('adasdasd') #快速生成集合
a = isdisjoint(b) #是否有共同元素
a = d.union(c) #将集合d和集合c合并生成集合a
a = d.difference(c) #返回只在d中不在c中的集合
a = d.intersection(c) #返回dc共同存在的集合
a.add(x)
a.remove(x)
a.discard(x) #如果x在a中，从a中移除
a.pop(x) #
##语句
print("I","LOVE","YOU",sep="❤") #用这个连接
x,y,z = 1,2,3 #同时赋值
key,value = d.popitem() #获取项，赋值，并从d中删除
a,b,*rest=[1,2,3,4] #将多余的值收集再rest中
x += 2 #增强赋值，x=x+2 *= 
bool() #转换为布尔值
#PY比较运算符 PY基础教程72
x is y #xy为同一对象
x is not y
x in y #x是容器y的成员
x not in y
x and y  #接受2个真值，两个都为真时返回真
x or y
x not y
assert a<b #错误时出现警报
#循环
while x <= 100:
    if x == 9:
        break  #当x为9时停止
while x <= 100:
    if x == 9:
        continue  #当x为9时返回到循环开头,跳入下一个循环
for num in []:    
for i in range(5):   #range(5)为迭代的次数，不需要显示在代码块中
while: #for循环后面可以加else 最后执行 ，除非之前加break
for i in word: #对于word中的每个字母
if 'u' not in word: #不在时执行for i in "aaa":
import time
time.sleep(5) #让程序暂停指定的秒数
#递归 一定要避免陷入死循环
#一个函数中调用了自己，这样的函数叫递归。
def getnum(num):
    if num > 1:
        return num*getnum(num-1)
    else:
        return num
    
#集合推导 从其他列表创建列表的方式,列表解析器
[x*x for x in range(5)] #[0, 1, 4, 9, 16] #后面循环为前面一共x值
[x*x for x in range(5) if x %3 == 0] #[0, 9]
#三人行
pass #什么也不做，可将其用作占位符，因为PY中的代码块不能为空
del x # 将x删除
exce("",scope) #将字符串作为代码执行，放置scope的空间中
eval() #计算用字符串表示的PY表达式的值，并返回结果。
##函数
#在一个项目中如果需要新函数，就在创建一个新函数，不要修改原先的代码，如果能用的旧函数
#那么直接调用旧函数，不要写新代码。
callable(x) #判断对象是否可以调用
def fun(A = "a"):    #定义新函数 提供默认值
    return fun #return用于从函数返回值！
#一个函数中可以有多个return，但是只有第一个返回给调用者，结束函数
return a,b,c #返回元祖
return [a,b,c] #返回列表
d = [a,b,c]
return d
def fun(a:str) -> set： #希望参数是字符串，并且返回集合

def fun():    #定义新函数
    'aaaaaaaaaa' #文档字符串,
    """
	aaaaaaa
	""" 
	#文档字符串，可以跨行
	
    return fun #return用于从函数返回值！
help(fun) #返回文档字符串
fun._doc_ #访问文档字符串
fun(A="a",B="b") #制定参数的名称
def fun(title,*params) #星号意味着收集余下的位置参数，如果没有，返回空元组。
#带星号的参数放在其他位置时，需要使用名称制定后续参数。
def fun(**params) #收集关键字参数，返回字典
global x #告诉 PY是全局变量
#列表和字典不用加global也可以当做全局变量，但是最好加，让其他人知道这是全局变量
#函数里用全局变量要加global，提醒是全局变量
#函数里不能直接修改全局变量，除非加global
#匿名函数 默认有return
def test(a,b,func): #匿名函数当实参
    result = func(a,b) 
    return result
num = test(11,22,lambda x,y:x+y)
###类
isinstance(x,A) #x是不是A的实例
delattr(a,'y') #删除a的y属性 相当于del a.y
getattr(a,'x') #获得a的x属性
hasattr(a,'x') #判断x是否为a的属性
setattr(a,'x',value) #新建
id() #查看标识码
#实例方法，self，只能通过对象名访问，不能通过类访问。
#静态方法，可用使用类名或对象名访问 @staticmethod
#类方法 第一个参数必须为cls, @classmethod
#控制访问 __name = '' 只能被类访问
dir(x) #查看所有属性名
vars(x) #以字典的形式返回x的所有属性名和属性值
__init__ #构造函数，他会在构造对象的时候自动调用，如果不设置函数会调用默认构造函数
#成员访问
A.x A.fun #都叫类的成员 函数可以外部定义 fun为方法
class fun(): #类中的函数称为方法
    def __init__(): #开头和结尾都有下划线。每次创建新实例PY都会自动运行他
#继承
issubclass(x,y) #x是不是继承了y
#只要子类中有初始化函数，就不会调用父类的初始化函数。__init__(self)
__slot__ #限定类的属性
__call__ #表示是否可调用
callable() #查看函数是否可以调用
class.__bases__ #知道基类
class A(B): #A继承B
    def __init__(self,...):
        super().__init__(...) #将父类和子类关联起来
from file import fun #从file.py中导入fun类
from file #导入整个模块
a =  file.fun() #调用类
from file1 import A
from file2 import B #若A的使用需要B的调用，分别导入即可

##创建发布文件
#1 names.py,setup.py,README.txt 都再mudel文件中
#2 按住shift 右键鼠标 再此处打开命令窗口，输入：
py -3 setup.py sdist
#3 shift按dist，右键鼠标，再此处打开命令窗口，输入：
py -3 -m pip install name.zip #dist中的压缩包，
#升级pip python -m pip install -U --force-reinstall pip


##储存和管理数据（文件操作）
name = open('names.txt','a') #采用追加模式
name = open('names.txt','r') #打开一个文件读取数据 ，默认
name = open('names.txt','w') #打开一个文件来写数据
name = open('names.txt','x') #打开一个新文件来写数据

file_path = 'F:\PY\《Python编程》源代码文件\chapter_10\pi_digits.txt'
with open(file_path) as file_object: #with打开的文件只在with代码块中使用
    contents = file_object.read()
    print(contents.rstrip())
    
with open(file_path) as file_object:
    for line in file_object：
		print(line) #逐行打印

with open(file_path，'w') as file_object: #若指定文件已经存在，PY将清空该文件
    file_object.write("I love programming.\n") #写入空文件
    file_object.write("I love YOU.\n") #写入空文件
    
name.close() #关闭文件流进行清理,with时不用close
file_object.read(5) #一次读5个字节
f.write("object")
f.write("\nobject")
##复制文件夹中的代码
old_file = open(old_file_path,"r")
new_file = open(new_file,'w')
content = old_file.read()
new_file.write(content)
old_file.close()
new_file.close()
#读取大文件
while True:
    content = old_file.read(1024)
    if len(content) == 0:
        break
    new_file.write(cotent)
#打开图片
from PIL import Image
img = Image.open('name')
img #在py中打开
img.show() #在windows中打开
#文件写入
import os
od.chdir('path') #声明工作目录
f = open('filename') #直接打开
f.read()
f.write() #写入
f.close() #关闭
f.readline() #读取一行
f.readlines() #读取所有文件，每一行一个元素，存在列表

for i in f.readlines():
    print(i) #读取多行

a = open('name').read()
print(a)
#定位读写
#seek第二个参数0表示文件的开头，1表示当前的位置，2表示末尾,第一个参数+向右-向左调
f = open(filename)
f.seek(2,0) #从文件的开头跳两个字节 使用seek可以重新读取 ，
f.tell() #当前的位置
##文件的常用操作
import os
os.rename('oldname.txt','filename.txt') #重命名
os.remove('name.txt') #删除文件
os.mkdir('file') #创建文件夹
os.rmdir('file') #删除文件夹
os.getwd() #返回绝对路径
os.chdir('') #改变默认路径
os.list("./") #获取当前路径下的目录列表
##批量重命名
import os
folder_name = "file" #输入文件夹中的名称
file_names =os.listdir(folder_name) #列出文件夹中所有文件的名字
for name in file_names:
    print(name)
    old_file_name = folder_name+"/"+name
    new_file_name = folder_name+"/"+"[OK]-"+name #新名字
    os.rename(old_file_name,new_file_name)
#pickle储存Python的原生对象
import pickle
D = [2,3,4]
F = open('name.pkl','wb')
pickle.dump(D,F)
F.close()

F = open('name.pkl','rb')
E = pickle.load(F) #读取
 
#打开CSV
import csv
filename = 'name'
with open(filename) as f:
    reader = csv.reader(f)
    header_row = next(reader) #返回文件中的下一行
    print(header)
    
    for index,column_header in enumerate(header_row): #获取每个元素的索引和值
        print(index,column_header)

    highs = [] 
    for row in reader: #遍历余下各行
        high = int(row[1])
        highs.append(high) #索引第二列的数据附加到末尾
    print(highs)

c,v = np.loadtxt('data.csv',delimiter=',',usecols=(6,7),unpack=True)  #numpy读取第六第7列
#异常
try:
    code #执行的代码
except Error: #输入异常对象
    code #输入出错时的对象
    pass #出错时继续运行，什么都不要做
else:
    code #正常运行时
##json
#json.loads() 把Json格式字符串解码转换成Python对象
#如果传入的字符串的编码不是UTF-8的话，需要指定字符编码的参数 encoding
dataDict = json.loads(jsonStrGBK)
dataDict = json.loads(jsonStrGBK, encoding="GBK")#改为这个
import json
strList = '[1, 2, 3, 4]'
strDict = '{"city": "北京", "name": "大猫"}'
json.loads(strList)  #[1, 2, 3, 4]
json.loads(strDict)  #{u'city': u'\u5317\u4eac', u'name': u'\u5927\u732b'}
#json.dumps() python类型转化为json字符串
import json
import chardet #chardet是一个非常优秀的编码识别模块
listStr = [1, 2, 3, 4]
tupleStr = (1, 2, 3, 4)
dictStr = {"city": "北京", "name": "大猫"}
json.dumps(listStr)  # '[1, 2, 3, 4]'
json.dumps(tupleStr) # '[1, 2, 3, 4]'
json.dumps(dictStr)  # '{"city": "\\u5317\\u4eac", "name": "\\u5927\\u5218"}'
chardet.detect(json.dumps(dictStr))# {'confidence': 1.0, 'encoding': 'ascii'}
print json.dumps(dictStr, ensure_ascii=False) # {"city": "北京", "name": "大刘"} #ensure_ascii=False有中文时
chardet.detect(json.dumps(dictStr, ensure_ascii=False))# {'confidence': 0.99, 'encoding': 'utf-8'}
#json.dump() Python内置类型序列化为json对象后写入文件
listStr = [{"city": "北京"}, {"name": "大刘"}]
json.dump(listStr, open("listStr.json","w"), ensure_ascii=False)
dictStr = {"city": "北京", "name": "大刘"}
json.dump(dictStr, open("dictStr.json","w"), ensure_ascii=False)
#json.load() 读取文件中json形式的字符串元素 转化成python类型
import json
strList = json.load(open("listStr.json"))
#储存
import json
filename = 'file.json'
with open(filename,'w') as a:
    json.dumps(objects)  #要储存的数据，可用于存数数据的文件filename

with open('dataname.json','w') as file: #用这个
    file.write(json.dumps(data))

with open('dataname1.json','w') as file: #若有中文字
    file.write(json.dumps(listStr,ensure_ascii=False))
##JsonPath JsonPath 对于 JSON 来说，相当于 XPATH 对于 XML。
XPath	JSONPath		描述
/			$			根节点
.			@			现行节点
/			.or[]		取子节点
..			n/a			取父节点，Jsonpath未支持
//			..			就是不管位置，选择所有符合条件的条件
*			*			匹配所有元素节点
@			n/a			根据属性访问，Json不支持，因为Json是个Key-value递归结构，不需要。
[]			[]			迭代器标示（可以在里边做简单的迭代操作，如数组下标，根据内容选值等）
|			[,]			支持迭代器中做多选。
[]			?()			支持过滤操作.
n/a			()			支持表达式计算
()			n/a			分组，JsonPath不支持

jsonobj = json.loads(html) #把json格式字符串转换成python对象
citylist = jsonpath.jsonpath(jsonobj,'$..name') #从根节点开始，匹配name节点
print(citylist)
fp = open('city.json','w')
content = json.dumps(citylist, ensure_ascii=False)
fp.write(content.encode('utf-8'))
fp.close()

##测试代码
import unittest
from model import fun #要测试的函数
class TestA(unittest.TestCase):
    '''测试medel.py中的fun函数'''
    def test_fun(self):
        '''是否能够测试'''
        test1 = fun()
        self.assertEqual(test1,"XX") #断言方法
unittest.main()

#断言方法
assertEqual(a,b)      #核实a==b
assertNotEqual(a,b)   #核实a!=b
assertTrue(x)         #核实x为True
assertFalse(x)        #核实x为False
assertIn(item,list)   #合适item在list中
assertNotIn(item,list)#合适item不在list中

####内建类型
type() #查看对象的类
bin() #二进制
oct() #八进制
hex() #16进制
##迭代类型
__iter__ #返回迭代器本身
__next__ #如果是第一次调用，则会返回对应数据容器中的第一个元素，否则返回上一次返回元素的下一个
#可用通过生成器快速构建可迭代对象，通过函数来定义，返回yield,生成一个结果，再进行迭代
def creator():
    L=[2,3,4]
    a=2
    b=3
    yield a
    yield b
    yield a+b
    yield from L
[2,3,5,2,3,4]

#迭代器 #迭代器仅仅在迭代到某个元素时才使用该元素,节省内存

from collections import Iterable
isinstance(object,Iterable)    #判断是否可以迭代
isinstance(object,Iterator)    #判断是否迭代对象，生成器一定是迭代器

a = iter(a) #将可迭代对象转换成迭代器
next(a) 

##闭包
#在函数内部在定义一个函数，并且这个函数用到了外面函数的变量，那么将这个函数以及用的一些
#变量成为闭包
#简化代码写作，若多次调用，虽然还是一个函数，但是创建一个对象保存新调用的函数

#生成器，保存一个算法，并不立即执行
#将列表[]变为()
b = (x*2 for x in range(10))
next(b) #每次取下一个
yield #只要加yield函数 就会变成生成器，将来调用包含yield的函数，会生成一个生成器对象
next(a) #会在yield会停止 并返回yield后面的值
next(a) #再一次执行会从yield下一次开始执行，循环到下一个yield

#迭代器可以用for循环
b.__next__() 和 next(b) #等价

temp = yield i  #等式右边 停止并返回i 。temp =   继续进行 temp没有赋值
b.send("hh") #将"hh"赋值给temp 不能第一次就调用 send中使用参数  要么第一次使用传入None
#多任务 看上去同时执行的任务 协程 进程 线程
############正则表达式
#match 方法：从起始位置开始查找，一次匹配 可传入起始位置
#search 方法：从任何位置开始查找，一次匹配 可传入起始位置
#findall 方法：全部匹配，返回列表 可传入起始位置
#finditer 方法：全部匹配，返回迭代器 可传入起始位置
#split 方法：分割字符串，返回列表 可传入起始位置
#sub 方法：替换
import re
r_object = re.compile(r'\d\d\d-\d\d\d-\d\d\d\d') #创建一个regex对象
match_oject = r_object.search('my number is 456-456-3211.') #创建一个匹配对象
match_oject = r_object.search('my number is 456-456-3211.',2,10) #从第3个位置到第11个位置
match_oject.group()  #返回匹配结果
match_oject.start()  #匹配对象的起始位置
match_oject.end()  #匹配对象的结束位置
match_oject.span()  #匹配对象的跨度，返回元组

r_object = re.compile(r'(\d\d\d-\d\d\d)-(\d\d\d\d)') #利用括号分组
match_oject = r_object.search('my number is 456-456-3211.') #创建一个匹配对象
match_oject.group(1)  #456-456
match_oject.group()   #456-456-3211
match_oject.groups()  #('456-456', '3211')

r_object = re.compile(r'\d\d\d-\d\d\d-\d\d\d\d') #利用括号分组
match_oject = r_object.findall('my number is 456-456-3211 and 142-546-7896') #返回一个字符串列表 
match_oject #['456-456-3211', '142-546-7896']

r_object = re.compile(r'(\d\d\d-\d\d\d)-(\d\d\d\d)')
match_oject = r_object.findall('my number is 456-456-3211 and 142-546-7896') #返回一个字符串的元组的列表
match_oject #[('456-456', '3211'), ('142-546', '7896')]

re.compile(r'[jzp.]']) #在方括号，普通的正则表达式不会被解释 不用加转义
re.compile(r'[^jzp]']) #非方括号内的字符
re.compile(r'^[jzp]']) #j、z、p开头的
re.compile(r'[jzp]$']) #j、z、p结尾的
#点-星组合
r_object = re.compile(r'First Name: (.*) Last Name:(.*)') #利用括号分组
match_oject = r_object.search('First Name: Ji Last Name:Zhipeng') #创建一个匹配对象
match_oject.group(1)  #Ji
match_oject.group(2)  #Zhipeng
r_object = re.compile(r'First Name: (.*?) Last Name:(.*?)') #.*?非贪心组合
#匹配换行
r_object = re.compile(r'(\d\d\d-\d\d\d)-(\d\d\d\d)',re.DOTALL) #匹配多行

r_object = re.compile(r'(jzp)',re.I) #不区分大小写
#spilt split 方法按照能够匹配的子串将字符串分割后返回列表
i = 'sad564asd54asd54asd2q4we53w'
p = re.compile(r'[\d]+')
print(p.split(i)) #['sad', 'asd', 'asd', 'asd', 'q', 'we', 'w']

##替换字符串sub(repl, string,count) count为替代的次数
r_object = re.compile(r'jzp \w+')
sub_ob = r_object.sub('JZP','jzp is GOOD!') #'JZP GOOD!'

reg_num = re.compile(r'(\d{3})\d{4}')
reg_num.sub(r'\1****','my num is 15898931538') #'my num is 158****1538'
#使用分组
import re
p = re.compile(r'(\w+) (\w+)') # \w = [A-Za-z0-9]
s = 'hello 123, hello 456'
print(p.sub(r'\2 ', s)) #第二组(\w+) 123 , 456 
#若是函数
def func(m):
    return 'hi' + ' ' + m.group(2)
p.sub(func, s) 
#忽略表达式中的空白符和注释
r_object = re.compile(r'''(\w+) #第一组
						  (\d+)	#第二组
						  (\s+)''',re.VERBOSE)  
#匹配中文
import re
title = u'你好，hello，世界'
pattern = re.compile(r'[\u4e00-\u9fa5]+')
result = pattern.findall(title)
print(result)
############处理Excel电子表格
import openpyxl
from openpyxl import Workbook
wb = Workbook() #新建一个excel对象
ws = wb.active #打开活动页 
ws['A1'] = 42
ws.append([1, 2, 3]) #下一行添加
ws['A3'] = datetime.datetime.now()
wb.save("sample1.xlsx") #将excel对象保存为excel

xlsx_object = openpyxl.load_workbook('example.xlsx')
xlsx_object.get_sheet_names() #得到sheet名称
xlsx_sheet1 = xlsx_object.get_sheet_by_name('sheet1')
sheet1 = xlsx_object.worksheets[0] #用这个
xlsx_sheet2 = xlsx_object.get_active_sheet() #得到打开时的活动表
sheet2 = xlsx_object.active #用这个
sheet1['A1'].value #a1的值
sheet1['A1'].row #单元格的行名
sheet1['A1'].column #单元格的列名
sheet1['A1'].coordinate #单元格的整体位置
sheet.cell(row=2,column=1) #定位单元格

############python与MYSQL
import pymysql
db = pymysql.connect(host='localhost',user='root',password = '',port=3306) #host为IP，user为用户名，pass为密码，port为端口
cursor = db.cursor() #获取操作游标，利用游标执行SQL语句
cursor.execute('select version()') #执行
data = cursor.fetchone() #获取第一条数据
print('Database version:',data)
cursor.execute('create database spiders default character set utf8')
db.close()
#创建表
db = pymysql.connect(host='localhost',user='root',password = '',port=3306) #host为IP，user为用户名，pass为密码，port为端口
cursor = db.cursor()
sql1 = 'use testdb' 
sql2 = 'CREATE TABLE IF NOT EXISTS students(id varchar(255) NOT NULL ,name varchar(255) NOT NULL,age INT NOT NULL,PRIMARY KEY(id))'
cursor.execute(sql1) #执行sql语句
cursor.execute(sql2) #执行sql语句
db.close() #关闭
#插入数据
import pymysql
id = '4'
user = 'Bob22'
age = 201
db = pymysql.connect(host='localhost',user='root',password = '',port=3306,db = 'testdb') #host为IP，user为用户名，pass为密码，port为端口
cursor = db.cursor()
sql = 'INSERT INTO students(id,name,age) values(%s,%s,%s)'
try:
    cursor.execute(sql,(id,user,age,))
    db.commit()
except:
    db.rollback()
db.close()
#标准写法
try:
    cursor.execute(sql)
    db.commit()
except:
    db.rollback()
#构造字典
db = pymysql.connect(host='localhost',user='root',password = '',port=3306,db = 'testdb') #host为IP，user为用户名，pass为密码，port为端口
data = {
    'id':'20120001',
    'name':'Bob',
    'age':20
    }
table = 'students'
keys = ','.join(data.keys())
values = ','.join(['%s']*len(data))
sql = 'INSERT INTO {table}({keys}) VALUES ({values})'.format(table = table,keys = keys,values = values)
try:
    if cursor.execute(sql,tuple(data.values())):
        print('OK')
        db.commit()
except:
    print('NOT OK')
    db.rollback()
db.close()
#更新数据
sql = 'UPDATA students SET age = %s WHERE name = %s'
try:
    cursor.execute(sql,(25,'BOB'))
    db.commit()
except:
    db.rollback()
db.close()
#查询数据
cursor.roacount() #查询将结果的条数
cursor.fetchone() #查询第一条结果
cursor.fetchall() #查询所有结果，不显示之前查找的


############python与MongoDB
#链接MongDB
import pymongo
client = pymongo.MongoClient(host = 'localhost',port = 27017) #地址host,端口port
#指定数据库
db = client.test1
db = client['test']
#指定集合
collection_stus = db.students
collection_stus = db['students']
#插入数据
student1 = {
    'id' : '20170101',
    'name' : 'OK',
    'age' : 12
}
student2 = {
    'id' : '201710101',
    'name' : 'OK',
    'age' : 12
}
result = collection_stus.insert_one(student2)
print(result.inserted_id) #调用插入数据的_id
result1 = collection_stus.insert_many([student1,student2])
print(result1.inserted_ids) #调用多个数据的_ids
#查询
result = collection_stus.find_one({'id':'20170101'})
print(result)
from bson.objectid import ObjectId
result = collection_stus.find_one({'_id':ObjectId("5b53d4dfaacdce227c43178c")}) #使用ID
print(result)
results = collection_stus.find({'age':20})
for result in results:
    print(result)
result = collection_stus.find({'age':{'$gt':18}}) #年龄大于20
$lt   #小于
$gt   #大于
$lte  #小于等于
$gte  #大于等于
$ne   #不等于
$in   #在范围内
{'age':{'$in':[20,23]}}
$nin  #不在范围内
{'age':{'$nin':[20,23]}}
result = collection_stus({'name':{'$regex':'^M.*'}}) #正则表达式
$regex   #正则表达式
$exists  #属性是否存在
$type    #类型判断 整数型还是字符串型
$mod     #数字模查询
$text    #文本查询
$where   #高级条件查询
#计数
count = collection_stus.find().count()
print(count)
count = collection_stus.find({'age':20}).count()
print(count)
#排序
results = collection_stus.find().sort('name',pymongo.ASCENDING) #升序 pymongo.DESCENDING降序
pirnt(result['name'] for result in results)
#偏移
result = collection_stus.find().sort('name',pymongo.ASCENDING).skip(2) #得到第三个及以后的元素
pirnt(result['name'] for result in results)
result = collection_stus.find().sort('name',pymongo.ASCENDING).skip(2).limit() #得到第三个及以后的前两个元素
pirnt(result['name'] for result in results)                       
#更新
condition = {'name':'jizhipeng'}
student = collection_stus.find_one(condition)
student['age'] = 25
result = collection_stus.update(condition,student)
print(result) #更改jizhipeng的年龄

result = collection_stus.update(condition,{'$set':student})
result = collection_stus.update_one(condition,{'$set':student})
print(result.matched_count,result.modified_count) #返回匹配的数据条数和影响的数据条数
#删除
result = collection_stus.remove({
    'name':'Kevin'
})
print(result)
result = collection_stus.delete_one({
    'name':'Kevin'
})
print(result.delete_one_count)
result = collection_stus.delete_many({
    'name':'Kevin'
})
print(result.delete_one_count)













